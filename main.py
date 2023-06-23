
from PyQt5.uic import loadUi
import sys
import openpyxl
from PyQt5.uic import loadUi
from openpyxl import Workbook, load_workbook
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import *
import resources
# from viewRecordWindow import viewRecordClass

class mainWindowClass(QMainWindow):
    def __init__(self):
        super(mainWindowClass, self).__init__()
        loadUi('mainWindow.ui', self)
        self.addRecordButton.clicked.connect(self.openSecondWindow) 
        self.viewRecordButton.clicked.connect(self.openViewRecordWindow)

    def openSecondWindow(self):
        self.newRecordWindow = newRecordWindow()
        self.newRecordWindow.show()

    def openViewRecordWindow(self):
        self.viewRecordClass = viewRecordClass()
        self.viewRecordClass.show()


class newRecordWindow(QMainWindow):
    def __init__(self):
        super(newRecordWindow, self).__init__()
        loadUi('newRecordWindow.ui', self)
        self.saveRecordButton.clicked.connect(self.saveData)
        self.clearPushButton.clicked.connect(self.clearValues)  
        self.cancelPushButton.clicked.connect(self.close)

    def saveData(self):
        start_time = self.startTimeEdit.time().toString("hh:mm:ss")
        end_time = self.endTimeEdit.time().toString("hh:mm:ss")
        battery_before = self.batteryBeforeDoubleSpinBox.value()
        battery_after = self.batteryAfterDoubleSpinBox.value()
        battery_used = battery_before - battery_after
        lights_on = self.lightsOnCheckBox.isChecked()
        distance = self.distanceDoubleSpinBox.value()
        max_speed = self.maxSpeedSpinBox.value()
        avg_speed = self.avgSpeedSpinBox.value()
        weight = self.weightSpinBox.value()
        speed_mode = self.speedModeSpinBox.value()
        average=distance/battery_used

        # Open the Excel file or create a new one if it doesn't exist
        try:
            workbook = load_workbook('data.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Start Time","End Time","Battery % Before","Battery % After","Speed Mode","Battery Used","Lights on","Distance(KM)","Max Speed(kmph)", "Avg. Speed(kmph) ", "Weight(Kg)","Average"])

        # Append a new row with the user input data
        sheet.append([start_time,end_time,battery_before,battery_after,speed_mode,battery_used,lights_on,distance,max_speed ,avg_speed,weight,average])

        # Save the changes to the Excel file
        workbook.save('data.xlsx')
        self.close()

    def clearValues(self):
        self.startTimeEdit.setTime(QtCore.QTime())  # Reset start time
        self.endTimeEdit.setTime(QtCore.QTime())  # Reset end time
        self.batteryBeforeDoubleSpinBox.setValue(0.0)  # Reset battery before
        self.batteryAfterDoubleSpinBox.setValue(0.0)  # Reset battery after
        self.distanceDoubleSpinBox.setValue(0.0)  # Reset distance
        self.maxSpeedSpinBox.setValue(0)  # Reset max speed
        self.avgSpeedSpinBox.setValue(0)  # Reset average speed
        self.speedModeSpinBox.setValue(0)  # Reset speed mode
        self.weightSpinBox.setValue(0)  # Reset weight
        self.lightsOnCheckBox.setChecked(False)  # Uncheck lights on checkbox

    def closeEvent(self, event):
        event.accept()

class tableWidgetClass(QMainWindow):
    def __init__(self):
        super(tableWidgetClass,self).__init__()
        loadUi('table.ui', self)  
        self.setWindowTitle("Available Records")
        self.load_data()     
        self.setEditable(False)  # Set the table as non-editable by default
        self.tableWidget.itemChanged.connect(self.save_data)  # Connect the itemChanged signal to the save_data method

    def flags(self, index: QtCore.QModelIndex) -> QtCore.Qt.ItemFlags:
        flags = super().flags(index)
        if not self._editable:
            flags = flags &~ QtCore.Qt.ItemIsEditable
        return flags

    def setEditable(self, editable):
        self._editable = editable 
        self.tableWidget.setEditTriggers(QtWidgets.QAbstractItemView.DoubleClicked if editable else QtWidgets.QAbstractItemView.NoEditTriggers)

    def load_data(self):
        path="data.xlsx"
        workbook=openpyxl.load_workbook(path)
        sheet = workbook.active

        self.tableWidget.setRowCount(sheet.max_row)
        self.tableWidget.setColumnCount(sheet.max_column)

        list_values=list(sheet.values)
        self.tableWidget.setHorizontalHeaderLabels(list_values[0])

        row_idx=0
        for value_tuple in list_values[1:]:
            col_idx=0
            for value in value_tuple:
                self.tableWidget.setItem(row_idx ,col_idx , QTableWidgetItem(str(value)))
                col_idx+=1
            row_idx+=1
        # self.tableWidget.setItem(0,2,QTableWidgetItem("hello"))
    
    def save_data(self, item):
        path = "data.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        row = item.row()
        col = item.column()
        value = item.text()

        sheet.cell(row=row + 2, column=col + 1).value = value  # Add 2 to row to account for header row

        workbook.save(path)

    def _removeRow(self):
        if self.tableWidget.rowCount() > 0:
            currentRow = self.tableWidget.currentRow()
            self.tableWidget.removeRow(currentRow)

        path = "data.xlsx"
        workbook = load_workbook(path)
        sheet = workbook.active

        # Remove the corresponding row from the sheet
        sheet.delete_rows(currentRow + 2, 1)  # Add 2 to currentRow to account for header row

        workbook.save(path)

    def resizeEvent(self, event):
        # Call the base class implementation
        super(tableWidgetClass, self).resizeEvent(event)

        # Adjust the size of the table widget to match the window size
        table_widget = self.tableWidget
        table_widget.setGeometry(0, 0, self.centralwidget.width(), self.centralwidget.height())

    

class viewRecordClass(QWidget):
    def __init__(self):
        super().__init__()
        self.resize(1600, 600)

        mainLayout = QHBoxLayout()
        table = tableWidgetClass()
        mainLayout.addWidget(table)
        buttonLayout = QVBoxLayout()

        button_new = QPushButton('New')
        button_new.clicked.connect(self.openSecondWindow)
        buttonLayout.addWidget(button_new)

        button_edit = QPushButton('Edit')
        button_edit.clicked.connect(lambda: table.setEditable(True))
        buttonLayout.addWidget(button_edit)

        button_remove = QPushButton('Remove')
        button_remove.clicked.connect(table._removeRow)
        buttonLayout.addWidget(button_remove,alignment=Qt.AlignTop)

        
        # button_save = QPushButton('Save')
        # # button_save.clicked.connect(save_data)
        # buttonLayout.addWidget(button_save, alignment=Qt.AlignTop)

        mainLayout.addLayout(buttonLayout)
        self.setLayout(mainLayout)

    def openSecondWindow(self):
        self.newRecordWindow = newRecordWindow()
        self.newRecordWindow.show()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    mainWindow = mainWindowClass()
    mainWindow.show()
    sys.exit(app.exec_())


